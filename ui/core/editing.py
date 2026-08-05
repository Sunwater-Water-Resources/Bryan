"""Editing a sims list - into a NEW file, never the master.

The master workbook is formula-driven: ``Output file``, ``Input MCDF``,
``Inflow``, ``SQ file``, ``Log file`` and ``Basename`` are typically Excel
concatenation formulas, which is what makes "add a row and the paths fill
themselves in" work. openpyxl cannot write a formula *and* its cached value, so
saving such a workbook from Python turns every formula into an uncached one and
Bryan then reads blanks.

That is not theoretical - it is exactly the state
``TFD_SimsList_LongList_02.xlsx`` is in. So this module never writes the file it
read: ``save_as`` refuses the master path outright, and what it writes holds
literal values with a banner saying so.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from .grouping import GROUP_COLUMN, add_group_keys
from .paths import cell_text
from .runwriter import write_chunk_list


class EditError(Exception):
    """An edit that must not be saved."""


@dataclass(frozen=True)
class CellChange:
    row: int
    column: str
    before: str
    after: str


def set_cell(frame: pd.DataFrame, index, column, value) -> pd.DataFrame:
    """Assign one cell, widening the column's dtype if it has to.

    A sims-list column that is blank all the way down reads as float64, and
    pandas 3 raises rather than silently upcasting when a string is put into
    it - which is every edit to an empty ``Comment`` or ``Lake config``. The
    column is converted to object first so the edit lands.
    """
    if column not in frame.columns:
        frame[column] = pd.Series([None] * len(frame), index=frame.index,
                                  dtype=object)

    if value is not None and not pd.api.types.is_object_dtype(frame[column].dtype):
        try:
            frame.loc[index, column] = value
            return frame
        except (TypeError, ValueError):
            frame[column] = frame[column].astype(object)

    frame.loc[index, column] = value
    return frame


def diff_frames(original: pd.DataFrame, edited: pd.DataFrame) -> list:
    """Every cell that changed, for review before anything is written."""
    changes = []
    for column in edited.columns:
        if column not in original.columns:
            for index in edited.index:
                after = cell_text(edited.loc[index, column])
                if after:
                    changes.append(CellChange(int(index), column, "", after))
            continue
        for index in edited.index:
            if index not in original.index:
                continue
            before = cell_text(original.loc[index, column])
            after = cell_text(edited.loc[index, column])
            if before != after:
                changes.append(CellChange(int(index), column, before, after))
    return changes


def duplicate_rows(frame: pd.DataFrame, indices, overrides=None) -> pd.DataFrame:
    """Copy rows, optionally overriding some cells.

    New rows are copies rather than blanks on purpose: a blank row is missing
    every column pre-flight requires, and the real editing task is 'do this
    group again at a different GWL'.
    """
    overrides = overrides or {}
    if not len(indices):
        raise EditError("choose a row to copy first")

    copies = frame.loc[list(indices)].copy()
    for column, value in overrides.items():
        copies[column] = value
    out = pd.concat([frame, copies], ignore_index=True)
    return out


def add_group_column(frame: pd.DataFrame) -> pd.DataFrame:
    """Write the derived group keys into a real ``Group`` column."""
    out = frame.copy()
    out[GROUP_COLUMN] = add_group_keys(frame)
    return out


def assert_not_master(destination, master) -> None:
    destination, master = Path(destination), Path(master)
    try:
        same = destination.resolve() == master.resolve()
    except OSError:
        same = str(destination) == str(master)
    if same:
        raise EditError(
            f"{master.name} is the master list and is never written by the UI. "
            f"Saving it from Python would replace its formulas with literal "
            f"values and lose the cached results Bryan reads. Choose a new "
            f"filename, or edit the master in Excel."
        )


def default_save_name(master, when=None) -> str:
    """A new name beside the master: <stem>_<yyyymmdd>_01.xlsx."""
    import datetime as _datetime

    master = Path(master)
    stamp = (when or _datetime.date.today()).strftime("%Y%m%d")
    for number in range(1, 100):
        candidate = master.with_name(f"{master.stem}_{stamp}_{number:02d}.xlsx")
        if not candidate.exists():
            return candidate.name
    return f"{master.stem}_{stamp}_99.xlsx"


def save_as(sims, edited: pd.DataFrame, destination, *, overwrite=False) -> Path:
    """Write ``edited`` to a new workbook, preserving shape and sheet name.

    Goes through the same writer the run copies use, so extra columns, column
    order and the sheet name survive, blanks stay blank, and a value that looks
    like a formula stays a literal.
    """
    destination = Path(destination)
    assert_not_master(destination, sims.path)
    if destination.exists() and not overwrite:
        raise EditError(f"{destination.name} already exists - confirm to replace it")

    if list(edited.columns) != list(sims.frame.columns):
        _write_reshaped(sims, edited, destination)
        return destination

    changed = {index for index in edited.index
               if index in sims.frame.index
               and not edited.loc[index].equals(sims.frame.loc[index])}
    new_rows = [index for index in edited.index if index not in sims.frame.index]

    if not changed and not new_rows:
        write_chunk_list(sims.path, list(edited.index), destination,
                         source_row_numbers=False)
        _restore_include(sims, edited, destination)
        return destination

    _write_reshaped(sims, edited, destination)
    return destination


def _write_reshaped(sims, edited: pd.DataFrame, destination: Path) -> None:
    """Write a frame that no longer matches the source row for row."""
    from openpyxl import Workbook

    from .runwriter import _native, _set

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sims.sheet_name

    for position, name in enumerate(edited.columns, start=1):
        _set(sheet.cell(row=1, column=position), name)
    for offset, index in enumerate(edited.index, start=2):
        for position, name in enumerate(edited.columns, start=1):
            _set(sheet.cell(row=offset, column=position),
                 _native(edited.loc[index, name]))

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _restore_include(sims, edited, destination: Path) -> None:
    """``write_chunk_list`` forces Include to 'yes'; an edit must not.

    Selecting rows to run and setting Include in the file are different
    things, and only the run copies should force it.
    """
    if "Include" not in edited.columns:
        return
    from openpyxl import load_workbook

    from .runwriter import _native, _set

    workbook = load_workbook(destination)
    sheet = workbook.worksheets[0]
    column = list(edited.columns).index("Include") + 1
    for offset, index in enumerate(edited.index, start=2):
        _set(sheet.cell(row=offset, column=column),
             _native(edited.loc[index, "Include"]))
    workbook.save(destination)


SAVE_NOTE = (
    "The saved copy holds literal values, not formulas. Excel writes a "
    "formula's result when it saves; Python cannot, so anything derived - the "
    "output names and paths - is frozen as it reads today. Edit the master in "
    "Excel if you need the formulas kept."
)
