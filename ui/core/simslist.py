"""Reading a sims list so that what the UI shows is what Bryan will run.

``read_sims_list`` makes the identical call Main.py:37 makes -
``pd.read_excel(path, sheet_name=0)`` - and holds that frame **unmodified** as
the source of truth. Everything the UI derives (group keys, selection,
predicted log paths, completion state) is a sidecar keyed by frame index, never
a mutation of the frame. That makes "what you see is what runs" a structural
property rather than a promise.

The other job here is the formula audit. See ``audit_formulas``.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd


@dataclass(frozen=True)
class FormulaAudit:
    """Formula cells whose cached value Excel never stored.

    Sims lists are formula-driven: ``Output file``, ``Input MCDF``, ``Inflow``,
    ``SQ file``, ``Log file`` and ``Basename`` are typically concatenation
    formulas. openpyxl and pandas both read the value **Excel cached** when it
    last saved. A workbook rewritten by a Python tool has formulas but no
    cached values, and then pandas - and therefore Bryan - reads them as blank.

    This is not hypothetical. ``TFD_SimsList_LongList_02.xlsx`` in the Tinaroo
    project has 96 formula cells and zero cached values, so its first 24 rows
    read with no ``Output file`` and no input paths. Every other sims list in
    that project and in Callide is 100% cached.
    """

    formula_cells: int = 0
    uncached_cells: int = 0
    uncached_by_column: dict = field(default_factory=dict)   # name -> count
    uncached_rows: frozenset = frozenset()                   # 0-based frame index

    @property
    def is_damaged(self) -> bool:
        return self.uncached_cells > 0

    def affects(self, row_index: int) -> bool:
        return row_index in self.uncached_rows

    def describe(self) -> str:
        if not self.is_damaged:
            return "all formula results are cached"
        columns = ", ".join(
            f"{name} ({count})"
            for name, count in sorted(self.uncached_by_column.items(),
                                      key=lambda item: -item[1])
        )
        return (
            f"{self.uncached_cells} of {self.formula_cells} formula cells have "
            f"no cached value, across {len(self.uncached_rows)} row(s). "
            f"Columns: {columns}. Bryan reads these as blank. Fix: open the "
            f"workbook in Excel, press Ctrl+Alt+F9 to recalculate, and save."
        )


@dataclass(frozen=True)
class SimsList:
    """A sims list as Bryan will see it, plus what the UI needs to be careful."""

    path: Path
    sheet_name: str
    other_sheets: tuple
    frame: pd.DataFrame
    audit: FormulaAudit
    sha256: str
    mtime: float

    @property
    def columns(self) -> list:
        return list(self.frame.columns)

    def row(self, index: int) -> pd.Series:
        return self.frame.loc[index]


def audit_formulas(path) -> FormulaAudit:
    """Compare the value view of a workbook against its formula view.

    Mirrors ``swe2d/config/batch.py:193-220``, with one fix: that code tests
    ``value is None``, but ``TFD_SimsList_LongList_02.xlsx`` stores ``<v></v>``
    - an *empty* cached value, not a missing one - so the test has to be
    ``value in (None, "")`` or the damage reads as clean.
    """
    from openpyxl import load_workbook

    values_wb = load_workbook(path, data_only=True, read_only=True)
    formulas_wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws_values = values_wb.worksheets[0]
        ws_formulas = formulas_wb.worksheets[0]

        rows_v = list(ws_values.iter_rows(values_only=True))
        rows_f = list(ws_formulas.iter_rows(values_only=True))
        if not rows_v:
            return FormulaAudit()

        header = [str(cell).strip() if cell is not None else ""
                  for cell in rows_v[0]]

        formula_cells = 0
        uncached_by_column: dict = {}
        uncached_rows: set = set()

        for offset, (value_row, formula_row) in enumerate(zip(rows_v[1:], rows_f[1:])):
            for index, name in enumerate(header):
                formula = formula_row[index] if index < len(formula_row) else None
                if not _is_formula(formula):
                    continue
                formula_cells += 1
                value = value_row[index] if index < len(value_row) else None
                if value in (None, ""):
                    label = name or f"column {index + 1}"
                    uncached_by_column[label] = uncached_by_column.get(label, 0) + 1
                    uncached_rows.add(offset)   # 0-based, matching the frame

        return FormulaAudit(
            formula_cells=formula_cells,
            uncached_cells=sum(uncached_by_column.values()),
            uncached_by_column=uncached_by_column,
            uncached_rows=frozenset(uncached_rows),
        )
    finally:
        values_wb.close()
        formulas_wb.close()


def _is_formula(cell) -> bool:
    """openpyxl gives a str starting with '=' , or an ArrayFormula object."""
    if isinstance(cell, str):
        return cell.startswith("=")
    return type(cell).__name__ == "ArrayFormula"


def read_sims_list(path) -> SimsList:
    """Read a sims list exactly as Main.py:37 does, plus provenance and audit."""
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"simulation list not found: {path}")

    frame = pd.read_excel(path, sheet_name=0)

    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True)
    try:
        sheet_name = workbook.worksheets[0].title
        other_sheets = tuple(ws.title for ws in workbook.worksheets[1:])
    finally:
        workbook.close()

    return SimsList(
        path=path,
        sheet_name=sheet_name,
        other_sheets=other_sheets,
        frame=frame,
        audit=audit_formulas(path),
        sha256=hashlib.sha256(path.read_bytes()).hexdigest()[:16],
        mtime=path.stat().st_mtime,
    )
