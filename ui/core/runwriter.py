"""Writing a run folder: one sims list and one config per chunk.

The master workbook is never touched. Each chunk gets a fresh single-sheet
workbook holding only its own rows, with ``Include`` forced to 'yes' and every
formula **flattened to the value Excel cached**.

Why an openpyxl cell-value copy and not ``DataFrame.to_excel``:

- ``to_excel`` mangles duplicate headers ('Log file', 'Log file.1');
- it round-trips dtypes, so a Duration of 24 becomes 24.0 - and
  lib/URBSmodel.py:395 builds storm filenames as ``f'ari{aep}_E{ensemble}.{duration}'``;
- it writes an index column unless told not to, and loses the sheet name.

A ``.value`` copy from a ``data_only=True`` load is provably what pandas read,
because pandas reads through openpyxl with the same flag.

The critical detail is step 5 in ``write_chunk_list``: openpyxl turns a string
starting with '=' back into a FORMULA, with no cached value. That is exactly
how TFD_SimsList_LongList_02.xlsx came to have 96 uncached cells. Every written
cell is forced to a string type to make that impossible.
"""

from __future__ import annotations

import datetime as _datetime
import os
import platform
import sys
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from .config import SimsConfig
from .grouping import add_group_keys
from .paths import atomic_write_json, safe_run_id

RUN_ROOT_NAME = "_ui_runs"
GROUP_COLUMN = "Group"
SOURCE_ROW_COLUMN = "Source row"


class WriteError(Exception):
    """The run folder could not be written correctly."""


@dataclass(frozen=True)
class ChunkFiles:
    index: int
    sims_list: Path
    config: Path
    console_log: Path
    run_log: Path
    rows: tuple
    simulation_list_value: str   # what goes in the config, relative to project


@dataclass(frozen=True)
class RunFolder:
    run_id: str
    folder: Path
    chunks: tuple
    manifest: Path

    @property
    def console_logs(self) -> list:
        return [chunk.console_log for chunk in self.chunks]


def make_run_id(when=None, tag="") -> str:
    """A sortable, filesystem- and RunLog-safe run id."""
    when = when or _datetime.datetime.now()
    stamp = when.strftime("%Y%m%d-%H%M%S")
    candidate = f"{stamp}-{tag}" if tag else stamp
    return safe_run_id(candidate)


def write_run(sims, config: SimsConfig, plan, run_id=None, *,
              group_keys=True, test_runs=None) -> RunFolder:
    """Write the whole run folder and return where everything landed."""
    run_id = run_id or make_run_id()
    safe_run_id(run_id)
    folder = config.project_folder / RUN_ROOT_NAME / run_id
    folder.mkdir(parents=True, exist_ok=True)

    groups = add_group_keys(sims.frame) if group_keys else None
    chunks = []
    for chunk in plan.chunks:
        name = f"chunk_{chunk.index:02d}"
        sims_path = folder / f"{name}.xlsx"
        config_path = folder / f"{name}.json"

        write_chunk_list(
            sims.path, chunk.rows, sims_path,
            log_file_overrides=plan.log_file_overrides,
            group_values=groups,
        )

        # Relative to the project folder, with forward slashes. os.path.join
        # takes it on Windows, and lib/RunLog.py:29's substring replace is
        # separator-agnostic, so chunk_NN_log.csv lands in this folder.
        list_value = f"{RUN_ROOT_NAME}/{run_id}/{name}.xlsx"
        write_chunk_config(config, config_path, list_value, run_id, chunk.index,
                           test_runs=test_runs)

        chunks.append(ChunkFiles(
            index=chunk.index,
            sims_list=sims_path,
            config=config_path,
            console_log=folder / f"{name}_console.txt",
            run_log=folder / f"{name}_log.csv",
            rows=tuple(chunk.rows),
            simulation_list_value=list_value,
        ))

    manifest = folder / "manifest.json"
    atomic_write_json(manifest, _manifest(sims, config, plan, run_id, chunks))
    _write_launch_scripts(folder, config, chunks)

    return RunFolder(run_id=run_id, folder=folder, chunks=tuple(chunks),
                     manifest=manifest)


def write_chunk_list(source: Path, rows, dest: Path, *,
                     log_file_overrides=None, group_values=None,
                     source_row_numbers: bool = True) -> None:
    """Write the selected rows of ``source`` to ``dest`` as plain values.

    Cell values come from the **pandas frame**, not from openpyxl, so the chunk
    reads back with the same dtypes the master list has. That is not fussiness:
    ``URBSmodel.duration_string`` returns '120.0' for a float and '120' for an
    int, which changes every storm filename, and ``get_simulation_period``
    looks the duration up in the URBS config's JSON keys as a string - a float
    misses and silently falls back to twice the storm duration (:287).

    Only the header row is copied from openpyxl, because pandas renames
    duplicate column names ('Log file.1') and the sheet's own text is what
    Bryan should see.
    """
    from openpyxl import Workbook, load_workbook

    log_file_overrides = log_file_overrides or {}
    frame = pd.read_excel(source, sheet_name=0)

    source_wb = load_workbook(source, data_only=True, read_only=True)
    try:
        sheet = source_wb.worksheets[0]
        header_row = next(sheet.iter_rows(values_only=True), None)
        if header_row is None:
            raise WriteError(f"{source} has no rows")
        header = list(header_row)[:len(frame.columns)]
        header += [None] * (len(frame.columns) - len(header))
        names = [str(cell).strip() if cell is not None else "" for cell in header]
        title = sheet.title
    finally:
        source_wb.close()

    def column_of(name):
        return names.index(name) if name in names else None

    include_at = column_of("Include")
    log_at = column_of("Log file")

    out_wb = Workbook()
    out = out_wb.active
    out.title = title

    extra = []
    if group_values is not None:
        extra.append(GROUP_COLUMN)
    if source_row_numbers:
        extra.append(SOURCE_ROW_COLUMN)

    for position, value in enumerate(header + extra, start=1):
        _set(out.cell(row=1, column=position), value)

    for offset, frame_index in enumerate(rows, start=2):
        excel_row = frame_index + 2          # header is row 1, frame is 0-based
        if frame_index not in frame.index:
            raise WriteError(
                f"row {frame_index} is not in {source.name} - the sims list "
                f"changed since it was read"
            )
        values = [_native(value) for value in frame.loc[frame_index].tolist()]

        if include_at is not None:
            values[include_at] = "yes"       # Main.py:64 wants exactly this
        if log_at is not None and frame_index in log_file_overrides:
            override = log_file_overrides[frame_index]
            if not (override is None or (isinstance(override, float) and pd.isna(override))):
                values[log_at] = override

        if group_values is not None:
            values.append(str(group_values.get(frame_index, "")))
        if source_row_numbers:
            values.append(excel_row)

        for position, value in enumerate(values, start=1):
            _set(out.cell(row=offset, column=position), value)

    dest.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(dest)
    _verify(source, rows, dest, header_names=names, extra=extra)


def _native(value):
    """A pandas/numpy cell value openpyxl can write.

    NaN and pd.NA must become None, or openpyxl writes the string 'nan' and
    Bryan reads a path called 'nan'. numpy scalars become their Python
    equivalents, keeping int as int and float as float - which is the whole
    point (see write_chunk_list).
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    item = getattr(value, "item", None)
    if callable(item) and getattr(value, "shape", ()) == ():
        try:
            return value.item()
        except (AttributeError, ValueError):
            return value
    return value


def _set(cell, value) -> None:
    """Assign a value, never a formula.

    openpyxl treats any string starting with '=' as a formula and stores it
    with no cached value - which is exactly the damage this whole module
    exists to avoid. Forcing the data type keeps a literal literal.
    """
    if isinstance(value, str) and value.startswith("="):
        cell.value = value
        cell.data_type = "s"
        return
    cell.value = value
    if isinstance(value, (_datetime.datetime, _datetime.date)):
        cell.number_format = "yyyy-mm-dd"


def _verify(source: Path, rows, dest: Path, header_names, extra) -> None:
    """Read the written file back and prove it matches the source rows.

    A production check, not a test. It is the only mechanism that shows the
    run copy holds what the UI displayed - and it would have caught the
    uncached-formula damage the moment it was introduced.
    """
    written = pd.read_excel(dest, sheet_name=0)
    original = pd.read_excel(source, sheet_name=0)

    expected = original.loc[list(rows)].reset_index(drop=True)
    shared = [name for name in expected.columns
              if name in written.columns and name not in ("Include", "Log file")
              and name not in extra]

    if len(written) != len(expected):
        raise WriteError(
            f"{dest.name}: wrote {len(written)} rows, expected {len(expected)}"
        )

    # Values are compared numerically where both sides are numbers. An xlsx
    # cell cannot carry the int/float distinction - openpyxl writes 120.0 as
    # "120" and pandas reads it back as int - so a dtype shift is not
    # corruption and is reported separately by runplan.stringify_hazards,
    # which is where it can be shown to the user before they commit.
    for name in shared:
        left = expected[name].reset_index(drop=True)
        right = written[name].reset_index(drop=True)
        for position in range(len(left)):
            a, b = left.iloc[position], right.iloc[position]
            if _differs(a, b):
                raise WriteError(
                    f"{dest.name}: column {name!r} row {position + 1} came back "
                    f"as {b!r}, source had {a!r}. The run copy would not run "
                    f"what was selected."
                )

    if "Include" in written.columns:
        bad = [value for value in written["Include"] if value != "yes"]
        if bad:
            raise WriteError(
                f"{dest.name}: {len(bad)} row(s) have Include != 'yes'; "
                f"Main.py:64 would skip them"
            )


def _differs(a, b) -> bool:
    """Whether two cell values disagree in a way that matters.

    Numbers compare numerically. A numpy int64 is not a Python int (numpy 2),
    so an isinstance check against (int, float) misses and would report the
    harmless 120.0 -> 120 dtype shift as corruption.
    """
    a_na, b_na = pd.isna(a), pd.isna(b)
    if a_na and b_na:
        return False
    if a_na or b_na:
        return True

    a_number, b_number = _as_float(a), _as_float(b)
    if a_number is not None and b_number is not None:
        return abs(a_number - b_number) > 1e-9
    return str(a) != str(b)


def _as_float(value):
    if isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def write_chunk_config(config: SimsConfig, dest: Path, simulation_list_value: str,
                       run_id: str, chunk_index: int, *, test_runs=None) -> dict:
    """The sims_config.json for one chunk.

    Everything here is chosen so Main.py's own resolution is a provable no-op:

    - ``project_folder`` absolute and pointing at the REAL project folder -
      otherwise ``dirname(config)`` would make the run folder the project
      folder and the simulation_list join would break (Main.py:29-35);
    - ``simulation_list`` relative to it, so lib/RunLog.py:29 also drops
      chunk_NN_log.csv inside the run folder;
    - ``filepaths`` absolute, because ``os.path.join(folder, abspath)`` returns
      the absolute path unchanged (Main.py:49).
    """
    payload = dict(config.raw)
    payload["project_folder"] = str(config.project_folder)
    payload["simulation_list"] = simulation_list_value
    payload["filepaths"] = {key: str(value) for key, value in config.filepaths.items()}
    payload["test_runs"] = config.test_runs if test_runs is None else int(test_runs)
    payload.pop("test run", None)          # the manual's spelling; avoid confusion
    payload["_ui"] = {
        "source_config": str(config.config_path),
        "source_simulation_list": config.simulation_list_rel,
        "run_id": run_id,
        "chunk": chunk_index,
        "written": _datetime.datetime.now().isoformat(timespec="seconds"),
    }
    atomic_write_json(dest, payload)
    return payload


def _manifest(sims, config, plan, run_id, chunks) -> dict:
    return {
        "run_id": run_id,
        "written": _datetime.datetime.now().isoformat(timespec="seconds"),
        "source": {
            "config": str(config.config_path),
            "simulation_list": str(sims.path),
            "sha256": sims.sha256,
            "mtime": sims.mtime,
            "sheet": sims.sheet_name,
            "dropped_sheets": list(sims.other_sheets),
        },
        "environment": {
            "python": sys.version.split()[0],
            "pandas": pd.__version__,
            "platform": platform.platform(),
            "computer": os.environ.get("COMPUTERNAME", platform.node()),
        },
        "plan": {
            "chunks": [
                {
                    "index": chunk.index,
                    "rows": list(chunk.rows),
                    "sims_list": chunk.sims_list.name,
                    "config": chunk.config.name,
                    "simulation_list_value": chunk.simulation_list_value,
                }
                for chunk in chunks
            ],
            "estimated_minutes": plan.estimated_minutes,
            "log_renames": [list(pair) for pair in plan.log_renames],
            "hazards": [
                {"severity": h.severity, "code": h.code, "message": h.message}
                for h in plan.hazards
            ],
        },
        "note": (
            "Formulas in the source workbook were flattened to the values "
            "Excel had cached. These chunk workbooks hold literal values only."
        ),
    }


def _write_launch_scripts(folder: Path, config: SimsConfig, chunks) -> None:
    """Re-runnable command lines, so a run does not depend on the UI existing.

    Mirrors the existing per-model .bat convention, including the ``cd /D`` -
    sims-list paths resolve against the process CWD.
    """
    project = config.project_folder
    bat = ["@echo off", "rem Written by the Bryan run launcher UI.",
           "rem Edit BRYAN_PY / BRYAN_MAIN to match this machine.", "setlocal",
           f'cd /D "{project}"',
           'set "BRYAN_PY=python"',
           'set "BRYAN_MAIN=Main.py"', ""]
    sh = ["#!/usr/bin/env bash",
          "# Written by the Bryan run launcher UI.",
          "set -euo pipefail",
          f'cd "{project}"',
          'BRYAN_PY="${BRYAN_PY:-python3}"',
          'BRYAN_MAIN="${BRYAN_MAIN:-Main.py}"', ""]

    for chunk in chunks:
        relative = os.path.relpath(chunk.config, project).replace(os.sep, "/")
        bat.append(f'"%BRYAN_PY%" -u "%BRYAN_MAIN%" "{relative}"')
        sh.append(f'"$BRYAN_PY" -u "$BRYAN_MAIN" "{relative}"')

    bat.append("pause")
    (folder / "launch.bat").write_text("\r\n".join(bat), encoding="utf-8")
    script = folder / "launch.sh"
    script.write_text("\n".join(sh) + "\n", encoding="utf-8")
    try:
        script.chmod(0o755)
    except OSError:
        pass
