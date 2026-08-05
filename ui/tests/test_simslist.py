"""Reading a sims list, and spotting the uncached-formula damage."""

from __future__ import annotations

import pandas as pd
import pytest

from conftest import (CALLIDE_COLUMNS, TINAROO_COLUMNS, reservoir_row,
                      write_workbook)
from core.simslist import audit_formulas, read_sims_list


def test_frame_is_exactly_what_pandas_reads(tmp_path):
    """The UI must show what Bryan will run - the same call, unmodified."""
    path = write_workbook(tmp_path / "sims.xlsx", TINAROO_COLUMNS,
                          [reservoir_row(Duration=d) for d in (18, 24, 36)])
    sims = read_sims_list(path)
    pd.testing.assert_frame_equal(sims.frame, pd.read_excel(path, sheet_name=0))


def test_sheet_name_and_other_sheets_are_recorded(tmp_path):
    path = write_workbook(tmp_path / "sims.xlsx", CALLIDE_COLUMNS,
                          [reservoir_row()], sheet_name="SIMS",
                          extra_sheets=("Definitions",))
    sims = read_sims_list(path)
    assert sims.sheet_name == "SIMS"
    assert sims.other_sheets == ("Definitions",)


def test_clean_workbook_reports_no_damage(tmp_path):
    path = write_workbook(tmp_path / "sims.xlsx", TINAROO_COLUMNS,
                          [reservoir_row()])
    audit = audit_formulas(path)
    assert not audit.is_damaged
    assert audit.uncached_cells == 0


def test_uncached_formulas_are_found_and_attributed(tmp_path):
    """Reproduces TFD_SimsList_LongList_02.xlsx: formulas, no cached values."""
    rows = [reservoir_row(Duration=d) for d in (18, 24)]
    formulas = {
        (offset, name): f'="x"&{name[0]}{offset + 2}'
        for offset in range(2)
        for name in ("Output file", "Input MCDF", "Inflow", "SQ file")
    }
    path = write_workbook(tmp_path / "damaged.xlsx", TINAROO_COLUMNS, rows,
                          formulas=formulas, cached=False)

    audit = audit_formulas(path)
    assert audit.is_damaged
    assert audit.formula_cells == 8
    assert audit.uncached_cells == 8
    assert audit.uncached_rows == frozenset({0, 1})
    assert set(audit.uncached_by_column) == {
        "Output file", "Input MCDF", "Inflow", "SQ file"}
    assert "Ctrl+Alt+F9" in audit.describe()


def test_uncached_cells_read_back_as_blank(tmp_path):
    """The reason it matters: Bryan sees nothing in those cells."""
    formulas = {(0, "Output file"): '="TFD_"&C2'}
    path = write_workbook(tmp_path / "damaged.xlsx", TINAROO_COLUMNS,
                          [reservoir_row()], formulas=formulas, cached=False)
    frame = pd.read_excel(path, sheet_name=0)
    assert pd.isna(frame.loc[0, "Output file"])


def test_empty_string_cached_value_counts_as_uncached(tmp_path):
    """The fix over swe2d's version.

    TFD_SimsList_LongList_02.xlsx stores <v></v> - an empty cached value, not a
    missing one - so a ``value is None`` test reports the damage as clean.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "Output file"
    cell = sheet.cell(row=2, column=1)
    cell.value = '="a"&"b"'
    path = tmp_path / "empty_cache.xlsx"
    workbook.save(path)

    # Rewrite the sheet XML so the formula carries an EMPTY cached value.
    import re
    import shutil
    import zipfile

    source = path.with_suffix(".orig.xlsx")
    shutil.move(path, source)
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(path, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith("sheet1.xml"):
                text = data.decode("utf-8")
                text = re.sub(r"(</f>)(?!<v>)", r"\1<v></v>", text)
                data = text.encode("utf-8")
            zout.writestr(item, data)

    audit = audit_formulas(path)
    assert audit.uncached_cells == 1, "an empty <v> must count as uncached"


def test_missing_file_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        read_sims_list(tmp_path / "nope.xlsx")
