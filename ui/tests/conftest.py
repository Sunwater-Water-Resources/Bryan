"""Fixtures for the UI tests.

Sims lists are **generated**, never committed: Bryan's .gitignore excludes
``*.xlsx`` and ``*.csv``, so a committed fixture would silently vanish from a
clone. The factories below build the shapes that actually occur in the wild.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

UI_ROOT = Path(__file__).resolve().parents[1]
BRYAN_ROOT = UI_ROOT.parent
for path in (str(UI_ROOT), str(BRYAN_ROOT)):
    if path not in sys.path:
        sys.path.insert(0, path)


# The two real schemas, taken from the Tinaroo and Callide projects.
TINAROO_COLUMNS = [
    "Include", "Output suffix", "Duration", "GWL", "Method", "Output file",
    "Run models", "Analyse results", "Store hydrographs", "Input MCDF",
    "Inflow", "ELS file", "SQ file", "FSL", "Hydrographs folder",
    "Results folder", "Config file", "Log file", "Comment",
]

CALLIDE_COLUMNS = [
    "Include", "Method", "Output suffix", "Model ID", "Duration", "Run models",
    "Analyse results", "Analyse volumes", "Store hydrographs", "Mop up files",
    "EBF on", "D50 on", "GWL", "FSL", "Input MCDF", "Inflow", "ELS file",
    "SQ file", "Hydrographs folder", "Basename", "Log file", "Output file",
    "Lake config", "ADV", "ADV source", "Config file", "Comment",
    "Results folder",
]

MONTE_CARLO_COLUMNS = [
    "Include", "Method", "Duration", "Run models", "Analyse results",
    "Store hydrographs", "Mop up files", "GWL", "ADV", "Lake config",
    "Focal subcatchments", "IL", "CL", "Exclusions", "Replicates",
    "Replicate file", "Config file", "Log file", "Output file", "Comment",
]


def write_workbook(path, columns, rows, *, sheet_name="Sheet1",
                   extra_sheets=(), formulas=None, cached=True):
    """Build an xlsx.

    ``formulas`` maps (row offset, column name) -> formula text. When
    ``cached`` is False the formula is written with no cached value, which is
    what a workbook rewritten by a Python tool looks like - and what
    TFD_SimsList_LongList_02.xlsx actually is.

    Ordinary values go through ``runwriter._set``, so a fixture holding a
    string that starts with '=' stays a literal. Writing it the naive way makes
    openpyxl store it as an uncached formula and the whole row reads back
    blank - the very failure these tests exist to pin.
    """
    from openpyxl import Workbook

    from core.runwriter import _set

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for position, name in enumerate(columns, start=1):
        _set(sheet.cell(row=1, column=position), name)

    formulas = formulas or {}
    for offset, row in enumerate(rows):
        for position, name in enumerate(columns, start=1):
            cell = sheet.cell(row=offset + 2, column=position)
            if (offset, name) in formulas and not cached:
                cell.value = formulas[(offset, name)]   # deliberately uncached
                continue
            _set(cell, row.get(name))

    for name in extra_sheets:
        workbook.create_sheet(title=name).cell(row=1, column=1).value = "notes"

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return Path(path)


def reservoir_row(**overrides):
    row = {
        "Include": "yes",
        "Output suffix": "FR-4C",
        "Duration": 18,
        "GWL": 1.7,
        "Method": "reservoir routing",
        "Output file": "TFD_mc_18h_GWL1p7",
        "Run models": "yes",
        "Analyse results": "yes",
        "Store hydrographs": "no",
        "Input MCDF": r"inputs\mcdf.csv",
        "Inflow": r"inputs\inflows.csv",
        "ELS file": r"reservoir\dam.els",
        "SQ file": r"reservoir\dam.sq",
        "FSL": 670.42,
        "Hydrographs folder": r"sims_mc\hydrographs",
        "Results folder": r"sims_mc\results",
        "Config file": r"sims_mc\mc_config.json",
        "Log file": r"sims_mc\log\run.log",
        "Comment": "",
    }
    row.update(overrides)
    return row


def monte_carlo_row(**overrides):
    row = {
        "Include": "yes",
        "Method": "monte carlo",
        "Duration": 24,
        "Run models": "yes",
        "Analyse results": "yes",
        "Store hydrographs": "no",
        "Mop up files": "yes",
        "GWL": 1.3,
        "ADV": "fsv",
        "Lake config": "",
        "Focal subcatchments": r"model\focal.csv",
        "IL": 30,
        "CL": 2.5,
        "Exclusions": "",
        "Replicates": "",
        "Replicate file": "",
        "Config file": r"sims_mc\mc_config.json",
        "Log file": r"sims_mc\log\mc.log",
        "Output file": r"sims_mc\results\CLD_mc_24h",
        "Comment": "",
    }
    row.update(overrides)
    return row


@pytest.fixture
def project(tmp_path):
    """A minimal project folder with a sims_config.json and the input files."""

    def build(columns, rows, *, sheet_name="Sheet1", extra_sheets=(),
              formulas=None, cached=True, config_name="sims_config.json",
              list_name="SimsList.xlsx", make_inputs=True):
        folder = tmp_path / "project"
        folder.mkdir(exist_ok=True)
        write_workbook(folder / list_name, columns, rows,
                       sheet_name=sheet_name, extra_sheets=extra_sheets,
                       formulas=formulas, cached=cached)

        for name in ("model_config.json", "storm_config.json", "climate_config.json"):
            (folder / name).write_text("{}", encoding="utf-8")

        (folder / config_name).write_text(json.dumps({
            "simulation_list": list_name,
            "filepaths": {
                "model_config": "model_config.json",
                "storm_config": "storm_config.json",
                "climate_config": "climate_config.json",
            },
            "test_runs": 0,
        }), encoding="utf-8")

        if make_inputs:
            for row in rows:
                for key in ("Input MCDF", "Inflow", "ELS file", "SQ file",
                            "Config file", "Focal subcatchments"):
                    value = row.get(key)
                    if not value:
                        continue
                    target = folder / str(value).replace("\\", "/")
                    target.parent.mkdir(parents=True, exist_ok=True)
                    if not target.exists():
                        target.write_text("{}" if target.suffix == ".json" else "x",
                                          encoding="utf-8")
        return folder / config_name

    return build


@pytest.fixture
def fake_main():
    """Path to the stand-in for Bryan's Main.py."""
    return Path(__file__).parent / "fake_main.py"
